from utilities import *
import logging
import traceback
import os, sys
from time import time
import numpy as np
import pandas as pd

from sklearn.preprocessing import OrdinalEncoder
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import balanced_accuracy_score, accuracy_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class ModelNonTop200(object):

    def __init__(self, df, samples_threshold):
        """ Class initialization, logging set-up, checking input files """
        self.logger = logging.getLogger('territory_finder_application.' + __name__)
        self.logger.debug("Model Non Top 200 is initializing")
        # Traget variable, auxulary targets and service fields
        self.target = 'Last_Future_ship_to'  
        self.target_aux = ['Region_Last_Future_Ship_to','Last_Future_ship_to_Name']
        self.service = ['isTrain','isCoord']
        # This model process all non top 200 outlets
        self.X_y = df[df['Trade_Structure']!='TOP200']
        self.logger.debug(f"X_y shape {self.X_y.shape}")
        # Unbalanced dataset. Use class_weight='balanced'
        self.model = RandomForestClassifier(class_weight='balanced', n_estimators=40,
            random_state=42, n_jobs=None, warm_start=False)
        # List of all features
        self.features = \
            [column for column in self.X_y.columns \
             if column not in [self.target] + self.target_aux + self.service]
        self.logger.debug(f"features: {self.features}")
        self.cat_features = self.X_y[self.features].select_dtypes(include=['object']).columns  # Categorical
        self.num_features = self.X_y[self.features].select_dtypes(exclude=['object']).columns  # Numeric
        # Ship-to classes for excluding
        self.ships_to_exclude = self.get_ships_to_exclude(self.X_y, samples_threshold)

    def get_ships_to_exclude(self, df, threshold=2):
        """ Return a list of classes with samples less than threshold """
        if threshold < 2:
            threshold = 2
        ship_counts = df[~df[self.target].isna()].groupby(self.target).size().to_frame()
        ship_counts.reset_index(inplace=True)
        ship_counts.columns = [self.target,'Counts']
        
        return [str(item) for item in list(ship_counts[self.target][ship_counts['Counts']<threshold].values)]

    def fill_nan(self, df):
        """ Fill NaN values, not affecting the target variable """
        for name in self.cat_features:
            df[name].fillna('missing', inplace=True)
        for name in self.num_features:
            df[name].fillna(0, inplace=True)
        self.logger.debug(who_am_i())
        return df
    
    def drop_nan(self):
        pass
    
    def get_encoded(self):
        """ Ordinal encoding implementation """
        # Fill nan values
        self.X_y = self.fill_nan(self.X_y)
        # Training encoders on full dataset (X & y separately)
        X = self.X_y[self.features]
        self.ordinal_encoder_x = OrdinalEncoder().fit(X[self.cat_features])
        y = self.X_y[self.X_y['isTrain']==True][self.target]
        self.label_encoder_y = LabelEncoder().fit(y)
        
        # full training dataset
        training = self.X_y[self.X_y['isTrain']==True]
        init_rows = training.shape[0]
        self.logger.debug(f"Rows in initial training dataset: {training.shape[0]}")
        # Rows to remove from training
        training_excluded = training[training[self.target].isin(self.ships_to_exclude)]
        self.logger.debug(f"Rows that excluded from the training: {training_excluded.shape[0]}")
        # Remove rows to remove from training
        training = training[~training[self.target].isin(self.ships_to_exclude)]
        self.logger.debug(f"Rows in training dataset: {training.shape[0]}")
        assert init_rows == training.shape[0] + training_excluded.shape[0]
        self.training = training

        # full dataset for prediction
        self.preds = self.X_y[self.X_y['isTrain']==False]
        self.logger.debug(f"Rows in prediction dataset: {self.preds.shape[0]}")
        
        # Training-Validation split
        self.X_train, self.X_valid, self.y_train, self.y_valid = \
                train_test_split(self.training[self.features], self.training[self.target],
                    test_size=0.3, random_state=42, stratify=self.training[self.target])

        def X_transform(df):
            # Transform of features
            X = df[self.features]
            X_cat = self.ordinal_encoder_x.transform(X[self.cat_features])
            # Not transform
            X_num = X[self.num_features]
            # Объединить категорийные и числовые
            X_enc = np.hstack([X_cat, X_num])

            return X_enc
        
        self.X_training_enc = X_transform(training)
        self.X_preds_enc = X_transform(self.preds)
        self.X_train_enc = X_transform(self.X_train)
        self.X_valid_enc = X_transform(self.X_valid)
        
        # Transform self.target variable
        y = training[self.target]
        # y is a DataFrame, converting to 1D array
        self.y_training_enc = self.label_encoder_y.transform(y.values.ravel())
        self.y_train_enc = self.label_encoder_y.transform(self.y_train.values.ravel())
        self.y_valid_enc = self.label_encoder_y.transform(self.y_valid.values.ravel())

    def get_x_preds(self):
        return self.preds[self.features]

    def get_x_valid(self):
        return self.X_valid

    def find_top_3(self, X_valid):
        """ Define top 3 classes for each outlet without an answer """
        y_pred_proba = self.model.predict_proba(X_valid)
        self.proba = pd.DataFrame(data=y_pred_proba, columns=self.model.classes_)
        self.logger.debug(f"proba.shape {self.proba.shape}")

        def get_max_3_classes(row):
            """ Get n-classes array and return TOP 3 classes with maximal probability """
            ser = pd.Series(data=row.values, index=self.model.classes_)
            ser.sort_values(inplace=True, ascending=False)
            return ser[0:3].index[0],ser[0:3].values[0], \
                ser[0:3].index[1],ser[0:3].values[1], \
                ser[0:3].index[2],ser[0:3].values[2]

        self.proba['top_1_class'], self.proba['top_1_proba'], \
            self.proba['top_2_class'], self.proba['top_2_proba'], \
            self.proba['top_3_class'], self.proba['top_3_proba'] = zip(*self.proba.apply(get_max_3_classes, axis=1))
        # inversion is needed
        self.proba['top_1_class'] = self.label_encoder_y.inverse_transform(self.proba['top_1_class'].values.ravel())
        self.proba['top_2_class'] = self.label_encoder_y.inverse_transform(self.proba['top_2_class'].values.ravel())
        self.proba['top_3_class'] = self.label_encoder_y.inverse_transform(self.proba['top_3_class'].values.ravel())
        # Preserve only new columns with TOP-3 classes
        self.proba = self.proba.loc[:,'top_1_class':]
        
    def get_statistics(self, X_valid, y_valid):
        """ Print statistics """
        self.find_top_3(X_valid)
        # y_valid is encoded, inversion is needed
        self.proba['y_valid'] = self.label_encoder_y.inverse_transform(y_valid)
        self.proba['correct_1'] = self.proba.apply(lambda x: int(x.top_1_class==x.y_valid),axis=1)
        self.proba['correct_2'] = self.proba.apply(lambda x: int(x.top_2_class==x.y_valid),axis=1)
        self.proba['correct_3'] = self.proba.apply(lambda x: int(x.top_3_class==x.y_valid),axis=1)
        # Total predictions
        total = self.proba.shape[0]
        # Number of correct predictions by class
        corr_cl1 = self.proba[self.proba.top_1_class==self.proba.y_valid].shape[0]
        corr_cl2 = self.proba[self.proba.top_2_class==self.proba.y_valid].shape[0]
        corr_cl3 = self.proba[self.proba.top_3_class==self.proba.y_valid].shape[0]
        not_correct = total - (corr_cl1 + corr_cl2 + corr_cl3)
        self.logger.info(f"""
        Total predictions: {total}
        Correct: {corr_cl1 / total * 100:.1f}% ({corr_cl1})
        Predicted in the 2nd option: {corr_cl2 / total * 100:.2f}% ({corr_cl2})
        Predicted in the 3rd option: {corr_cl3 / total * 100:.3f}% ({corr_cl3})
        Not predicted at all {not_correct / total * 100:.3f}% ({not_correct})
        """)

        def get_proba_info(class_num, proba_from, proba_to):
            """ Get class and return a number of correct, incorrect predictions in interval """
            correct_num = self.proba[(self.proba[class_num+'_class']==self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            incorrect_num = self.proba[(self.proba[class_num+'_class']!=self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            return correct_num, incorrect_num, (proba_from, proba_to)

        correct, incorrect, index = [], [], []
        for edge in range(20,100,10):
            cor, inc, ind = get_proba_info('top_1',edge/100,(edge+10)/100)
            correct.append(cor)
            incorrect.append(inc)
            index.append(ind)

        # Print out ratio of wrong and right predictions in each interval            
        a = pd.DataFrame(data={'correct':correct[::-1], 'incorrect':incorrect[::-1]}, index=index[::-1])
        top = 1
        rep_list = []
        rep_list.append(f"\n{'Interval':>12} {'Right':>8} {'Wrong':>10} {'Wrn./Tot.':>11}\n")
        for i in range(len(a)):
            mid = a.index[i][0]
            s = f"{str(a.index[i]):>12} {a.correct[i]:>8} {a.incorrect[i]:>10}"
            if i==0:
                v = a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100
            else:
                v = a.incorrect[:i+1].sum() / (a.incorrect[:i+1].sum() + a.correct[:i+1].sum()) * 100
            rep_list.append("{0} {1:>10.2f} | interval ({2}, {3}] wrong / total = {4:.2f}%\n" \
                    .format(s, a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100, mid, top, v))
        self.logger.info(''.join(rep_list))
        
    def validate(self):
        """
        Split the training dataset into the training and the validation parts
        for training and validation
        
        """
        self.get_encoded()
        # Training, Validation, Cross-Validation
        t0 = time()
        # Cross-Validation goes on full training dataset
        val_cv_score = cross_val_score(self.model, self.X_training_enc, self.y_training_enc, cv=3, scoring='balanced_accuracy')
        self.val_cv_score = np.array([round(item, 5) for item in val_cv_score])
        self.logger.info(f"Cross-validation average score: {self.val_cv_score.mean():.3f}")
        self.logger.debug(f"Cross-validation finished in {time() - t0:.3f} sec.")
        t0 = time()
        self.model.fit(self.X_train_enc, self.y_train_enc)
        self.logger.debug(f"Training finished in {time() - t0:.3f} sec.")
        t0 = time()
        y_pred = self.model.predict(self.X_valid_enc)
        self.val_score = balanced_accuracy_score(self.y_valid_enc, y_pred)
        self.logger.info(f"Balanced accuracy score: {self.val_score:.3f}")
        self.logger.debug(f"Validation finished in {time() - t0:.1f} sec.")
        # print statistics
        self.get_statistics(self.X_valid_enc, self.y_valid_enc)

    def fit(self):
        """ Training on full data set """
        self.logger.info("Final training...")
        self.get_encoded()
        t0 = time()
        self.model.fit(self.X_training_enc, self.y_training_enc)
        self.logger.debug(f"Final training finished in {time() - t0:.3f} sec.")
        self.find_top_3(self.X_preds_enc)

        return self.proba

class ModelTop200(object):

    def __init__(self, df, samples_threshold):
        """ Class initialization, logging set-up, checking input files """
        self.logger = logging.getLogger('territory_finder_application.' + __name__)
        self.logger.debug("Model Top 200 is initializing")
        # Traget variable, auxulary targets and service fields
        self.target = 'Last_Future_ship_to'
        self.target_aux = ['Region_Last_Future_Ship_to','Last_Future_ship_to_Name']
        self.service = ['isTrain','isCoord']
        # This model process only top 200 outlets using reduced set of features
        self.X_y = df[(df['Trade_Structure']=='TOP200')][['Chain_Id','Kladr_level_1','Ship_To_Visited',
            'isTrain','Region_Last_Future_Ship_to','Last_Future_ship_to_Name','Last_Future_ship_to']]
        # The trainig dataset includes rows:
        # 1. with all NaN target (to prediction)
        # 2. or only target that equal to Ship_To_Visited (for training)
        # (i.e. to exclude rows where target is not equal to Ship_To_Visited)
        self.X_y = self.X_y[(self.X_y[self.target].isna()) | \
                            ((~self.X_y[self.target].isna()) & \
                            (self.X_y['Ship_To_Visited']==self.X_y[self.target]))]
        self.X_y_excluded = self.X_y[(~self.X_y[self.target].isna()) & \
                                    (self.X_y['Ship_To_Visited']!=self.X_y[self.target])]
        self.logger.debug(f"X_y shape {self.X_y.shape}")
        self.logger.debug(f"X_y_excluded shape {self.X_y_excluded.shape}")
        # Unbalanced dataset. Use class_weight='balanced'
        self.model = DecisionTreeClassifier(class_weight='balanced', random_state=42)        
        # List of features
        self.features = \
            [column for column in self.X_y.columns \
             if column not in [self.target] + self.target_aux + self.service]
        self.logger.debug(f"features: {self.features}")
        self.cat_features = self.X_y[self.features].select_dtypes(include=['object']).columns  # Categorical
        self.num_features = self.X_y[self.features].select_dtypes(exclude=['object']).columns  # Numeric
        # Ship-to classes for excluding
        self.ships_to_exclude = self.get_ships_to_exclude(self.X_y, samples_threshold)

    def get_ships_to_exclude(self, df, threshold=2):
        """ Return a list of classes with samples less than threshold """
        if df.empty:
            self.logger.error(f"Envoke the load_data method first!")
            raise Exception
        if threshold < 2:
            threshold = 2
        ship_counts = df[~df[self.target].isna()].groupby(self.target).size().to_frame()
        ship_counts.reset_index(inplace=True)
        ship_counts.columns = [self.target,'Counts']
        
        return [str(item) for item in list(ship_counts[self.target][ship_counts['Counts']<threshold].values)]

    def fill_nan(self, df):
        """ Fill NaN values, not affecting the target variable """
        for name in self.cat_features:
            df[name].fillna('missing', inplace=True)
        for name in self.num_features:
            df[name].fillna(0, inplace=True)
        self.logger.debug(who_am_i())

        return df
    
    def drop_nan(self):
        pass
    
    def get_encoded(self):
        """ Ordinal encoding implementation """
        # Fill nan values
        self.X_y = self.fill_nan(self.X_y)
        # Training encoders on full dataset (X & y separately)
        X = self.X_y[self.features]
        self.ordinal_encoder_x = OrdinalEncoder().fit(X[self.cat_features])
        y = self.X_y[self.X_y['isTrain']==True][self.target]
        self.label_encoder_y = LabelEncoder().fit(y)
        
        # full training dataset
        training = self.X_y[self.X_y['isTrain']==True]
        init_rows = training.shape[0]
        self.logger.debug(f"Rows in initial training dataset: {training.shape[0]}")
        # Rows to remove from training
        training_excluded = training[training[self.target].isin(self.ships_to_exclude)]
        self.logger.debug(f"Rows that excluded from the training: {training_excluded.shape[0]}")
        # Remove rows to remove from training
        training = training[~training[self.target].isin(self.ships_to_exclude)]
        self.logger.debug(f"Rows in training dataset: {training.shape[0]}")
        assert init_rows == training.shape[0] + training_excluded.shape[0]
        self.training = training

        # full dataset for prediction
        preds = self.X_y[self.X_y['isTrain']==False]
        init_rows = preds.shape[0]
        self.logger.debug(f"Rows in initial prediction dataset: {init_rows}")
        # Rows to remove from prediction (where Ship_To_visited is missed)
        preds_excluded = preds[preds['Ship_To_Visited']=='missing']
        self.logger.debug(f"Rows that excluded from the prediction: {preds_excluded.shape[0]}")
        # Remove rows from prediction
        preds = preds[preds['Ship_To_Visited']!='missing']        
        self.logger.debug(f"Rows in prediction dataset: {preds.shape[0]}")
        assert init_rows == preds.shape[0] + preds_excluded.shape[0]
        self.preds = preds

        # Training-Validation split
        self.X_train, self.X_valid, self.y_train, self.y_valid = \
                train_test_split(self.training[self.features], self.training[self.target],
                    test_size=0.3, random_state=42, stratify=self.training[self.target])
        
        def X_transform(df):
            # Transform of features
            X = df[self.features]
            X_cat = self.ordinal_encoder_x.transform(X[self.cat_features])
            # Not transform
            X_num = X[self.num_features]
            # Объединить категорийные и числовые
            X_enc = np.hstack([X_cat, X_num])

            return X_enc
        
        self.X_training_enc = X_transform(training)
        self.X_preds_enc = X_transform(self.preds)
        self.X_train_enc = X_transform(self.X_train)
        self.X_valid_enc = X_transform(self.X_valid)
        
        # Transform self.target variable
        y = training[self.target]
        # y is a DataFrame, converting to 1D array
        self.y_training_enc = self.label_encoder_y.transform(y.values.ravel())
        self.y_train_enc = self.label_encoder_y.transform(self.y_train.values.ravel())
        self.y_valid_enc = self.label_encoder_y.transform(self.y_valid.values.ravel())
        
    def get_x_preds(self):
        return self.preds[self.features] 

    def get_x_valid(self):
        return self.X_valid

    def find_top_3(self, X_valid):
        """ Define top 3 classes for each outlet without an answer """
        y_pred_proba = self.model.predict_proba(X_valid)
        self.proba = pd.DataFrame(data=y_pred_proba, columns=self.model.classes_)
        self.logger.debug(f"proba.shape {self.proba.shape}")

        def get_max_3_classes(row):
            """ Get n-classes array and return TOP 3 classes with maximal probability """
            ser = pd.Series(data=row.values, index=self.model.classes_)
            ser.sort_values(inplace=True, ascending=False)
            return ser[0:3].index[0],ser[0:3].values[0], \
                ser[0:3].index[1],ser[0:3].values[1], \
                ser[0:3].index[2],ser[0:3].values[2]

        self.proba['top_1_class'], self.proba['top_1_proba'], \
            self.proba['top_2_class'], self.proba['top_2_proba'], \
            self.proba['top_3_class'], self.proba['top_3_proba'] = zip(*self.proba.apply(get_max_3_classes, axis=1))
        # inversion is needed
        self.proba['top_1_class'] = self.label_encoder_y.inverse_transform(self.proba['top_1_class'].values.ravel())
        self.proba['top_2_class'] = self.label_encoder_y.inverse_transform(self.proba['top_2_class'].values.ravel())
        self.proba['top_3_class'] = self.label_encoder_y.inverse_transform(self.proba['top_3_class'].values.ravel())
        # Preserve only new columns with TOP-3 classes
        self.proba = self.proba.loc[:,'top_1_class':]
        
    def get_statistics(self, X_valid, y_valid):
        """ Print statistics """
        self.find_top_3(X_valid)
        # y_valid is encoded, inversion is needed
        self.proba['y_valid'] = self.label_encoder_y.inverse_transform(y_valid)
        self.proba['correct_1'] = self.proba.apply(lambda x: int(x.top_1_class==x.y_valid),axis=1)
        self.proba['correct_2'] = self.proba.apply(lambda x: int(x.top_2_class==x.y_valid),axis=1)
        self.proba['correct_3'] = self.proba.apply(lambda x: int(x.top_3_class==x.y_valid),axis=1)
        # Total predictions
        total = self.proba.shape[0]
        # Number of correct predictions by class
        corr_cl1 = self.proba[self.proba.top_1_class==self.proba.y_valid].shape[0]
        corr_cl2 = self.proba[self.proba.top_2_class==self.proba.y_valid].shape[0]
        corr_cl3 = self.proba[self.proba.top_3_class==self.proba.y_valid].shape[0]
        not_correct = total - (corr_cl1 + corr_cl2 + corr_cl3)
        self.logger.info(f"""
        Total predictions: {total}
        Correct: {corr_cl1 / total * 100:.1f}% ({corr_cl1})
        Predicted in the 2nd option: {corr_cl2 / total * 100:.2f}% ({corr_cl2})
        Predicted in the 3rd option: {corr_cl3 / total * 100:.3f}% ({corr_cl3})
        Not predicted at all {not_correct / total * 100:.3f}% ({not_correct})
        """)

        def get_proba_info(class_num, proba_from, proba_to):
            """ Get class and return a number of correct, incorrect predictions in interval """
            correct_num = self.proba[(self.proba[class_num+'_class']==self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            incorrect_num = self.proba[(self.proba[class_num+'_class']!=self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            return correct_num, incorrect_num, (proba_from, proba_to)

        correct, incorrect, index = [], [], []
        for edge in range(20,100,10):
            cor, inc, ind = get_proba_info('top_1',edge/100,(edge+10)/100)
            correct.append(cor)
            incorrect.append(inc)
            index.append(ind)

        # Print out ratio of wrong and right predictions in each interval            
        a = pd.DataFrame(data={'correct':correct[::-1], 'incorrect':incorrect[::-1]}, index=index[::-1])
        top = 1
        rep_list = []
        rep_list.append(f"\n{'Interval':>12} {'Right':>8} {'Wrong':>10} {'Wrn./Tot.':>11}\n")
        for i in range(len(a)):
            mid = a.index[i][0]
            s = f"{str(a.index[i]):>12} {a.correct[i]:>8} {a.incorrect[i]:>10}"
            if i==0:
                v = a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100
            else:
                v = a.incorrect[:i+1].sum() / (a.incorrect[:i+1].sum() + a.correct[:i+1].sum()) * 100
            rep_list.append("{0} {1:>10.2f} | interval ({2}, {3}] wrong / total = {4:.2f}%\n" \
                    .format(s, a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100, mid, top, v))
        self.logger.info(''.join(rep_list))
        
    def validate(self):
        """
        Split the training dataset into the training and the validation parts
        for training and validation
        
        """
        self.get_encoded()
        # Training, Validation, Cross-Validation
        t0 = time()
        # Cross-Validation goes on full training dataset
        val_cv_score = cross_val_score(self.model, self.X_training_enc, self.y_training_enc, cv=3, scoring='balanced_accuracy')
        self.val_cv_score = np.array([round(item, 5) for item in val_cv_score])
        self.logger.info(f"Cross-validation average score: {self.val_cv_score.mean():.3f}")
        self.logger.debug(f"Cross-validation finished in {time() - t0:.3f} sec.")
        t0 = time()
        self.model.fit(self.X_train_enc, self.y_train_enc)
        self.logger.debug(f"Training finished in {time() - t0:.3f} sec.")
        t0 = time()
        y_pred = self.model.predict(self.X_valid_enc)
        self.val_score = balanced_accuracy_score(self.y_valid_enc, y_pred)
        self.logger.info(f"Balanced accuracy score: {self.val_score:.3f}")
        self.logger.debug(f"Validation finished in {time() - t0:.1f} sec.")
        # print statistics
        self.get_statistics(self.X_valid_enc, self.y_valid_enc)

    def fit(self):
        """ Training on full data set """
        self.logger.info("Final training...")
        self.get_encoded()
        t0 = time()
        self.model.fit(self.X_training_enc, self.y_training_enc)
        self.logger.debug(f"Final training finished in {time() - t0:.3f} sec.")
        self.find_top_3(self.X_preds_enc)

        return self.proba

class TerritoryFinder(object):
    
    def __init__(self, coord_file, report_file, output_file, samples_threshold=2):
        """ Class initialization, logging set-up, checking input files """
        self.logger = logging.getLogger('territory_finder_application.' + __name__)
        # input and output files
        self.coord_file, self.report_file, self.output_file = coord_file, report_file, output_file
        # self.df = pd.DataFrame()
        # self.X_train_enc, self.y_enc_train, self.X_preds_enc = None, None, None
        self.check_files()
        # Порог для исключения классов
        self.samples_threshold = samples_threshold
        # Traget variable, auxulary targets and service fields
        self.target = 'Last_Future_ship_to'  
        self.target_aux = ['Region_Last_Future_Ship_to','Last_Future_ship_to_Name']
        self.service = ['isTrain','isCoord']
        self.logger.debug("TerritoryFinder class initialized")
    
    def check_files(self):
        """ Check if input files exist and output file is closed """
        try:
            # Check if input files exist
            text = ""
            if not os.path.isfile(self.coord_file):
                text += f"File '{self.coord_file}' not found. Please place it in a folder with this program " \
                    "and set a correct name in the command line\n"
            if not os.path.isfile(self.report_file):
                text += f"File '{self.report_file}' not found. Please place it in a folder with this program " \
                    "and set a correct name in the command line\n"
            if text != "":
                raise FileNotFoundError
            # Check if the output file is closed (can not rewrite if it is opened)
            f = open(self.output_file, "a")
            f.close()
        except FileNotFoundError as err:
            self.logger.exception(text)
            sys.exit(1)
        except PermissionError as err:
            self.logger.exception(f"File '{self.output_file}' is opened. Please close the file or use another name")
            sys.exit(1)
        self.logger.debug(f"Input files were found")

    def create_sample_files(self):
        """ Creates sample input files for developing purposes """
        # Remove dupes and non-active outlets, then save a fraction of data in 10%
        df_terr = pd.read_excel(self.report_file, skiprows=1)
        df_terr = df_terr[df_terr['Основная / Дубликат']!='Дубликат']
        df_terr = df_terr[df_terr['Store Status NOW']!='Неактивная']
        df_terr_sample = df_terr.sample(frac=0.1, random_state=42)
        df_terr_sample.shape
        with pd.ExcelWriter('Report Territory Management_sample.xlsx') as writer: # pylint: disable=abstract-class-instantiated
            df_terr_sample.to_excel(writer, sheet_name='Sheet1', index=False)
        # Leave only used outlets in coord file 
        df_coor = pd.read_excel(self.coord_file)
        df_coor_sample = df_coor[df_coor['OL_ID'].isin(df_terr_sample['SWE Store Key'])]
        with pd.ExcelWriter('Coordinates_sample.xlsx') as writer: # pylint: disable=abstract-class-instantiated
            df_coor_sample.to_excel(writer, sheet_name='Sheet1', index=False)
        
    def get_avg_coordinate(self, row, kladr_grouped):
        """
        Return average coordinate of locality

        Parameters:
        -----------
        row (Series): ['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4'] locality for getting coordinate
        kladr_grouped (Series with Multiindex (['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']):
            contains coordinate for 4-level locality

        Returns:
        --------
        float: coordinate

        """
        try:
            return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3'],row['Kladr_level_4']]
        except KeyError:
            try:
                return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3']].mean()
            except KeyError:
                try:
                    return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2']].mean()
                except KeyError:
                    try:
                        return kladr_grouped[row['Kladr_level_1']].mean()
                    except KeyError:
                        text = f"Cannot get average coordinate for the locality {row['Kladr_level_1']}, " \
                            f"{row['Kladr_level_2']}, {row['Kladr_level_3']}, {row['Kladr_level_4']}"
                        self.logger.debug(text)
                        return 0
    
    def restore_coordinate_part(self, part_name):
        try:
            t0 = time()
            # Series with MultiIndex
            kladr_grouped = self.df[self.df['isCoord']==1]. \
                groupby(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4'])[part_name].mean()
            self.df.loc[self.df['isCoord']==0,'part_name'] = \
                self.df.loc[self.df['isCoord']==0] \
                [['SWE_Store_Key','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']].apply( \
                self.get_avg_coordinate, args=(kladr_grouped,), axis=1)
        except Exception as err:
            self.logger.exception(err)
        finally:
            self.logger.debug(f"{part_name} restoring finished in {time() - t0:.3f} sec. " \
                f"(not found {self.df[self.df[part_name]==0].shape[0]})")

    def restore_coordinates(self):
        """ Find coordinates for an outlet by its neighbors """
        self.df['Latitude'].replace(0, np.NaN, inplace=True)
        self.df['Longitude'].replace(0, np.NaN, inplace=True)
        self.df['isCoord'] = ~( (self.df['Latitude'].isna()) | (self.df['Longitude'].isna()) )
        # Restore Latitude
        self.restore_coordinate_part('Latitude')
        # Restore Longitude
        self.restore_coordinate_part('Longitude')
        # remove unnecessary fields
        self.df.drop(['Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5'],
                     axis=1, inplace=True)
    
    def load_coordinates(self):
        """ Load and pre-process coordinates """
        self.logger.info(f"Loading coordinates...")
        df = pd.read_excel(self.coord_file)
        self.logger.debug(f"Rows in {self.coord_file}: {df.shape[0]}")
        df.columns = ['SWE_Store_Key','Latitude','Longitude']
        # cleansing of invalid coordinates
        df = df[df['Latitude']!=0]
        df = df[(df['Latitude']>40)&(df['Latitude']<82)]
        df = df[((df['Longitude']>=10)&(df['Longitude']<180)) | \
            ((df['Longitude']>=-180)&(df['Longitude']<-160))]
        # check if outlets are duplicated
        if df.SWE_Store_Key.value_counts().values[0] > 1:
            self.logger.warning(f"Found duplicated codes of outlets in '{self.coord_file}!")
        return df

    def align_value(self, value):
        """ Remove symbols from left/right and duped commas """
        try:
            aligned = value
            try:
                aligned = str(int(float(aligned)))
            except ValueError:
                aligned = aligned.strip().replace(', ,',',').replace(',  ,',',') \
                    .replace(',,',',').replace(',,',',').replace(',,',',')
                # From left
                while aligned[0] not in DIGITS:
                    aligned = aligned[1:]
                # From right
                while aligned[-1] not in DIGITS:
                    aligned = aligned[:-1]
                aligned = np.array(aligned.split(',')).astype('float').astype('int')
                aligned = ','.join(aligned.astype(str))
        except Exception as err:
            self.logger.warning(f"Cannot format the Last_Future_ship_to value {value}. Error {err}")
            return value
        else:
            return aligned

    def load_report(self):
        """ Load and pre-process report """
        self.logger.info(f"Loading report file...")
        df = pd.read_excel(self.report_file, skiprows=1)
        self.logger.debug(f"Rows in {self.report_file}: {df.shape[0]}")
        # rename fields
        df.columns = REPORT_FILE_ALL_COLUMNS
        # Save a full list of outlet codes to use when formatting a new report
        self.df_codes = pd.DataFrame(data=df['SWE_Store_Key'],columns=['SWE_Store_Key'])
        # Remove outlet-duplicates and associated fields
        number_before = df.shape[0]
        df = df[df['General_Duplicate']!='Дубликат']
        df.drop(['Covering_Outlet_id','General_Duplicate'], axis=1, inplace=True)
        self.logger.debug(f"Removed duped outlets: {number_before - df.shape[0]}")
        # Remove inactive outlets
        number_before = df.shape[0]
        df = df[df['Store_Status_NOW']!='Неактивная']
        self.logger.debug(f"Removed inactive outlets: {number_before - df.shape[0]}")
        # Remove unused columns
        df = df[REPORT_FILE_COLUMNS]
        # Remove with NaN SWE key
        number_before = df.shape[0]
        df = df[~df['SWE_Store_Key'].isna()]
        if number_before - df.shape[0] > 0:
            self.logger.warning(f"Removed with NaN SWE key: {number_before - df.shape[0]}")
        # Mark outlet for training
        df['isTrain'] = ~ df[self.target].isna()
        # Align columns with classes
        df.loc[df['isTrain']==True,self.target] = \
            df.loc[df['isTrain']==True][self.target].apply(self.align_value)
        # df.loc[~df['Ship_To_Visited'].isna(),'Ship_To_Visited'] = \
        #     df.loc[~df['Ship_To_Visited'].isna()]['Ship_To_Visited'].apply(self.align_value)
        df['From_Dc'] = df['From_Dc'].astype(int)
        df['Chain_Id'] = df['Chain_Id'].astype(float).astype(int)
        return df

    def load_data(self):          
        """ Load and transform data """
        df_coor = self.load_coordinates()
        df_terr = self.load_report()
        # Merging territories with coordinates
        self.df = pd.merge(df_terr, df_coor, on='SWE_Store_Key',how='left')
        # del self.df_terr
        # del self.df_coor
        self.logger.info("Restore coordinates...")
        self.restore_coordinates()
        # Set SWE_Store_Key as index thereby exclude it from features
        self.df.set_index('SWE_Store_Key',inplace=True)

    def validate(self):
        model = ModelNonTop200(self.df, self.samples_threshold)
        model.validate()
        model_top200 = ModelTop200(self.df, self.samples_threshold)
        model_top200.validate()

    def get_report(self):
        """ Prepare a new report """
        self.logger.info("Calculate proba...")
        t0 = time()
        # Generate proba for non TOP 200 model
        model1 = ModelNonTop200(self.df, self.samples_threshold)
        proba1 = model1.fit()
        # Generate proba for TOP 200 model
        model2 = ModelTop200(self.df, self.samples_threshold)
        proba2 = model2.fit()
        
        X_pred1 = model1.get_x_preds()
        X_pred1.reset_index(inplace=True)
        df1 = pd.concat([X_pred1['SWE_Store_Key'], proba1], axis=1,join='inner')
        
        X_pred2 = model2.get_x_preds()
        X_pred2.reset_index(inplace=True)
        df2 = pd.concat([X_pred2['SWE_Store_Key'], proba2], axis=1,join='inner')
        
        df = pd.concat([df1, df2])

        df_info = self.df_codes.merge(right=df, how='left', on='SWE_Store_Key')
        df_info['SWE_Store_Key'] = df_info['SWE_Store_Key'].astype('str')
        self.logger.debug(f"Done in {time() - t0:.3f} sec.")
        
        target_aux_values = self.df[self.df['isTrain']==True][['Region_Last_Future_Ship_to','Last_Future_ship_to_Name',self.target]] \
            .groupby(self.target)['Region_Last_Future_Ship_to','Last_Future_ship_to_Name'].first()

        self.logger.info("Open report...")
        t0 = time()
        self.workbook = openpyxl.load_workbook(self.report_file)
        self.logger.debug(f"Done in {time() - t0:.3f} sec.")
        
        self.logger.info("Format report...")
        t0 = time()
        worksheet = self.workbook['Sheet1']
        rows = dataframe_to_rows(df_info, index=False, header=True)
        # Row and columns for text print out
        row_proba = 2
        col_proba = 54    # BB column
        col_region, col_name, col_ship_to = 49, 50, 51   # AY column
        # Setup proba columns width, title text, font, and alignment
        proba_widths = [19,11,5,11,5,11,5]
        proba_captions = ['SWE Store Key','1 class',' 1 proba','2 class','2 proba','3 class','3 proba']
        title_font = Font(name='Times New Roman', size=12, bold=True)
        title_alignment = Alignment(horizontal='left', vertical='top')
        title_fill = PatternFill("solid", fgColor="00CCFFCC")
        for i in range(7):
            worksheet.column_dimensions[get_column_letter(col_proba + i)].width = proba_widths[i]
            cell = get_column_letter(col_proba + i) + str(row_proba)
            worksheet[cell] = proba_captions[i]
            worksheet[cell].font = title_font
            worksheet[cell].alignment = title_alignment
            worksheet[cell].fill = title_fill
        # Define style for proba rows
        row_font = Font(name='Arial', size=10, bold=False)
        row_alignment = Alignment(horizontal='left', vertical='top')
        # Go through all rows
        for r_idx, row in enumerate(rows, row_proba):
            # If proba is defined
            if type(row[2])==float and not pd.isnull(row[2]):
                c1_ship_to, c1_proba = row[1], float(row[2])
                c1_region, c1_name = target_aux_values.loc[c1_ship_to][0], target_aux_values.loc[c1_ship_to][1]
                # Print and format Last Future columns
                if c1_proba >= 0.9:
                    fgColor = "00CCFFCC"
                elif c1_proba >= 0.7:
                    fgColor = "00FFFF99"
                else:
                    fgColor = "00FF9900"
                row_fill = PatternFill("solid", fgColor=fgColor)
                worksheet.cell(row=r_idx, column=col_region, value=c1_region).fill = row_fill
                worksheet.cell(row=r_idx, column=col_name, value=c1_name).fill = row_fill
                worksheet.cell(row=r_idx, column=col_ship_to, value=c1_ship_to).fill = row_fill
                # Print and format additional proba columns
                for c_idx, value in enumerate(row, col_proba):
                    worksheet.cell(row=r_idx, column=c_idx, value=value).font = row_font
                    worksheet.cell(row=r_idx, column=c_idx).alignment = row_alignment
                    worksheet.cell(row=r_idx, column=c_idx).fill = row_fill
                    
        self.logger.debug(f"Done in {time() - t0:.3f} sec.")

    def save_report(self):
        self.logger.info("Save output file...")
        t0 = time()
        self.workbook.save(self.output_file)
        self.logger.debug(f"Saved in {time() - t0:.3f} sec.")
        self.logger.info(f"New report saved as '{self.output_file}''")