import logging
import os, sys
from time import time
import numpy as np
import pandas as pd

from sklearn.preprocessing import OrdinalEncoder
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import balanced_accuracy_score
from sklearn.ensemble import RandomForestClassifier
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class TerritoryFinder(object):

    def __init__(self, coord_file, report_file, output_file, samples_threshold=2):
        """ Class initialization, logging set-up, checking input files """
        # input and output files
        self.coord_file, self.report_file, self.output_file = coord_file, report_file, output_file
        # Выборка не сбалансированная, используем class_weight='balanced', n_estimators=40
        self.model = RandomForestClassifier(class_weight='balanced', n_estimators=40, random_state=42, n_jobs=-1, warm_start=False)
        self.df = pd.DataFrame()
        self.X_enc_train, self.y_enc_train, self.X_enc_pred = None, None, None
        self.set_up_logging()
        self.check_env()
        # Порог для исключения классов
        self.samples_threshold = samples_threshold
        self.log.debug("TerritoryFinder class initialized")

    def set_up_logging(self):
        """ Set up logging """
        os.makedirs('logs', exist_ok=True)
        self.log = logging.getLogger(__name__)
        self.log.setLevel(logging.DEBUG)
        formatter = logging.Formatter(fmt="%(asctime)s %(levelname)s: %(message)s", datefmt="%Y-%m-%d - %H:%M:%S")
        sh = logging.StreamHandler(sys.stdout)
        sh.setLevel(logging.DEBUG)
        sh.setFormatter(formatter)
        fh = logging.FileHandler(u"./logs/territory_finder.log", "w")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(formatter)
        self.log.addHandler(sh)
        self.log.addHandler(fh)
    
    def check_env(self):
        if not os.path.isfile(self.coord_file):
            self.log.error(f"File '{self.coord_file}' not found. Please place it in a folder with this program")
            raise Exception
        if not os.path.isfile(self.report_file):
            self.log.error(f"File '{self.report_file}' not found. Please place it in a folder with this program")
            raise Exception
        self.log.debug(f"Input files were found")

    def restore_coordinates(self):
        """ Find coordinates for an outlet by its neighbors """

        self.df['Latitude'].replace(0, np.NaN, inplace=True)
        self.df['Longitude'].replace(0, np.NaN, inplace=True)
        self.df['isCoord'] = ~( (self.df['Latitude'].isna()) | (self.df['Longitude'].isna()) )

        kladr_lat_grouped = self.df[self.df['isCoord']==1].groupby(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']).Latitude.mean()
        kladr_lon_grouped = self.df[self.df['isCoord']==1].groupby(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']).Longitude.mean()

        def get_avg_coordinate(row, kladr_grouped):
            """
            Вернуть среднюю координату населенного пункта, области, региона или страны. Используем функции
            multiindex.isin().any(), чтобы проверить, что в Series имеется индекс для всех 4-х уровней
            и вернуть значение. В случае отсутствия индекса, отрубить последний уровень в индексе и проверить
            индекс для 3-х уровней и т.д.

            Parameters:
            row (Series): ['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4'] для которых нужно
                получить координату
            kladr_grouped (Series): с мультииндексом (['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4'])
                который содержит значения координаты для 4-х уровней из адресного классификатора
                
            Returns:
            float: Координата

            """
            try:
                if kladr_grouped.index \
                        .isin([(row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3'],row['Kladr_level_4'])]).any():
                    return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3'],row['Kladr_level_4']]
                elif kladr_grouped.index.droplevel(['Kladr_level_4']) \
                        .isin([(row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3'])]).any():
                    return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3']].mean()
                elif kladr_grouped.index.droplevel(['Kladr_level_3','Kladr_level_4']) \
                        .isin([(row['Kladr_level_1'],row['Kladr_level_2'])]).any():
                    return kladr_grouped[row['Kladr_level_1'],row['Kladr_level_2']].mean()
                elif kladr_grouped.index.droplevel(['Kladr_level_2','Kladr_level_3','Kladr_level_4']) \
                        .isin([(row['Kladr_level_1'])]).any():
                    return kladr_grouped[row['Kladr_level_1']].mean()
                else:
                    return 0
            except:
                print(row['Kladr_level_1'],row['Kladr_level_2'],row['Kladr_level_3'],row['Kladr_level_4'])
                raise KeyError

        self.df.loc[self.df['isCoord']==0,'Latitude'] = \
            self.df.loc[self.df['isCoord']==0][['SWE_Store_Key','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']].apply( \
                get_avg_coordinate, args=(kladr_lat_grouped,), axis=1)

        self.df.loc[self.df['isCoord']==0,'Longitude'] = \
            self.df.loc[self.df['isCoord']==0][['SWE_Store_Key','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']].apply( \
                get_avg_coordinate, args=(kladr_lon_grouped,), axis=1)

        self.df.drop(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5'], axis=1, inplace=True)

    def load_data(self):          
        """ Load and transform data """

        self.log.info(f"Loading coordinates...")
        df_coor = pd.read_excel(self.coord_file, nrows=1000)
        self.log.debug(f"Rows in {self.coord_file}: {df_coor.shape[0]}")
        df_coor.columns = ['SWE_Store_Key','Latitude','Longitude']
        # cleansing from invalid coordinates
        df_coor = df_coor[df_coor['Latitude']!=0]
        df_coor = df_coor[(df_coor['Latitude']>40)&(df_coor['Latitude']<82)]
        df_coor = df_coor[((df_coor['Longitude']>=10)&(df_coor['Longitude']<180)) | \
            ((df_coor['Longitude']>=-180)&(df_coor['Longitude']<-160))]

        # check if outlets are duplicated
        if df_coor.SWE_Store_Key.value_counts().values[0] > 1:
            self.log.error(f"Found duplicated codes of outlets in '{self.coord_file}!")
            raise Exception

        self.log.info(f"Loading report file...")
        df_terr = pd.read_excel(self.report_file, skiprows=1)
        self.log.debug(f"Rows in {self.report_file}: {df_terr.shape[0]}")
        # rename fields
        df_terr.columns = ['Region','Distrib','Office','FFDSL','TSE_MTDE',
            'Level_Torg_Region1','Level_Torg_Region2','Filial_Name',
            'Filial_Ship_To','Chain_Type','Chain_Name','Chain_Id',
            'Chain_Chain_Tier_MWC','Chain_Chain_Sub_Tier_MWC','SWE_Store_Key',
            'Store_Status','Store_Status_NOW','Outlet_Name','Channel_Name_2018',
            'Outlet_Type_2018','Trade_Structure','From_Dc',
            'Segment_MWC_Segment_Name','Cluster_MWC','Kladr_level_1',
            'Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5',
            'LSV_WWY','LSV_CHOCO','LSV_MWC','Covering_Outlet_id',
            'General_Duplicate','Ship_To_Visited','Filial_Visited',
            'Ship_to_Name_TO_BE','Region_loaded_RSS',
            'MW_Ship_to_TO_BE_Name_loaded_RSS',
            'MW_Ship_to_TO_BE_loaded_RSS',
            'CH_Ship_to_TO_BE_Name_loaded_RSS',
            'CH_Ship_to_TO_BE_loaded_RSS',
            'WR_Ship_to_TO_BE_Name_loaded_RSS',
            'WR_Ship_to_TO_BE_loaded_RSS','Ship_to_Code_TO_BE',
            'DC','Changed',
            'Change_Period','Region_Last_Future_Ship_to',
            'Last_Future_ship_to_Name', 'Last_Future_ship_to', 'Comment']

        self.df_codes = pd.DataFrame(data=df_terr['SWE_Store_Key'],columns=['SWE_Store_Key'])

        # do not take unused columns
        df_terr = df_terr[['SWE_Store_Key','Region','Distrib','Office','FFDSL','TSE_MTDE','Level_Torg_Region1',
            'Level_Torg_Region2','Filial_Name','Filial_Ship_To','Chain_Type','Chain_Id','Chain_Chain_Tier_MWC',
            'Chain_Chain_Sub_Tier_MWC','Channel_Name_2018','Outlet_Type_2018','Trade_Structure','From_Dc',
            'Segment_MWC_Segment_Name','Cluster_MWC','Covering_Outlet_id','General_Duplicate','Ship_To_Visited',
            'Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5',
            'Region_Last_Future_Ship_to','Last_Future_ship_to_Name','Last_Future_ship_to']]

        # Remove outlet-duplicates and associated fields
        df_terr = df_terr[df_terr['General_Duplicate']!='Дубликат']
        df_terr.drop(['Covering_Outlet_id','General_Duplicate'], axis=1, inplace=True)

        self.log.info("Merging territories with coordinates and start preprocessing...")
        self.df = pd.merge(df_terr, df_coor, on='SWE_Store_Key',how='left')
        del df_terr
        del df_coor

        self.log.info("Restore coordinates...")
        self.restore_coordinates()

        self.df['isTrain'] = ~ self.df['Last_Future_ship_to'].isna()

        # Last_Future_ship_to убрать внешние пробелы и преобразовать к типу str
        # self.df['Last_Future_ship_to'] = self.df['Last_Future_ship_to'].astype(str)
        self.df.loc[self.df['isTrain']==True,'Last_Future_ship_to'] = \
            self.df.loc[self.df['isTrain']==True]['Last_Future_ship_to'].apply(self.align_value)

        self.df['From_Dc'] = self.df['From_Dc'].astype(int)
        self.df['Chain_Id'] = self.df['Chain_Id'].astype(str)
        # Установить поле как индекс, тем самым исключив его из списка признаков
        self.df.set_index('SWE_Store_Key',inplace=True)
        # Классы для исключения
        self.ships_to_exclude = self.get_ships_to_exclude(self.samples_threshold)

    def align_value(self, value):
        """
        Избавиться от крайних символов и дублирующихся запятых.
        Если получен список, то отсортировать по возрастанию

        """
        try:
            aligned = value
            # Избавится от крайних символов и дублирующихся запятых
            try:
                aligned = str(int(float(aligned)))
            except ValueError:
                aligned = aligned.strip().replace(', ,',',').replace(',  ,',',') \
                    .replace(',,',',').replace(',,',',').replace(',,',',')
                while aligned[0] not in ['0','1','2','3','4','5','6','7','8','9']:
                    aligned = aligned[1:]
                while aligned[-1] not in ['0','1','2','3','4','5','6','7','8','9']:
                    aligned = aligned[:-1]
                aligned = np.array(aligned.split(',')).astype('float').astype('int')
#                 aligned = np.sort(aligned)
                aligned = ','.join(aligned.astype(str))
        except Exception as e:
            print(f"Возникла проблема при обработке кода {value}. Ошибка {e}")
            return value
        finally:
            return aligned

    def get_ships_to_exclude(self, threshold=2):
        """
        Вернуть список классов с количеством сэмплов меньшим threshold
    
        """
        if self.df.empty:
            self.log.error(f"Envoke the load_data method first!")
            raise Exception
        if threshold < 2:
            threshold = 2
        ship_counts = self.df[~self.df['Last_Future_ship_to'].isna()].groupby('Last_Future_ship_to').size().to_frame()
        ship_counts.reset_index(inplace=True)
        ship_counts.columns = ['Last_Future_ship_to','Counts']
        
        return [str(item) for item in list(ship_counts['Last_Future_ship_to'][ship_counts['Counts']<threshold].values)]

    def get_encoded(self):
        """ Ordinal encoding implementation """

        # fill NaN values, not touching target variable
        cat_features = self.df.select_dtypes(include=['object']).columns  # Categorical
        num_features = self.df.select_dtypes(exclude=['object']).columns  # Numeric
        self.df['Last_Future_ship_to'].replace(np.NaN, 0, inplace=True)
        for name in cat_features:
            self.df[name].fillna('missing', inplace=True)
        for name in num_features:
            self.df[name].fillna(0, inplace=True)
        self.df['Last_Future_ship_to'].replace(0, np.NaN, inplace=True)

        target = ['Last_Future_ship_to']         # Целевая переменная
        aux_target = ['Region_Last_Future_Ship_to','Last_Future_ship_to_Name']
        service = ['isTrain','isCoord']  # Сервисные признаки, которые отбросить
        # Из полного списка признаков или из переданного списка признаков исключить target, aux_target & service
        self.features = [column for column in self.df.columns \
                    if column not in target + aux_target + service]
        
        # Обучить кодировщик на полном наборе признаков
        X = self.df[(~self.df['Last_Future_ship_to'].isin(self.ships_to_exclude))][self.features]
        cat_features = X.select_dtypes(include=['object']).columns  # Categorical
        num_features = X.select_dtypes(exclude=['object']).columns  # Numeric
        self.encoder_x = OrdinalEncoder()
        self.encoder_x.fit(X[cat_features])

        X_train = self.df[(self.df['isTrain']==True)&(~self.df['Last_Future_ship_to'].isin(self.ships_to_exclude))][self.features]
        X_cat = self.encoder_x.transform(X_train[cat_features])       # Transform cats features
        X_num = X_train[num_features]                                   # Not transform nums features
        self.X_enc_train = np.hstack([X_cat, X_num])                  # Join

        X_pred = self.df[(self.df['isTrain']==False)&(~self.df['Last_Future_ship_to'].isin(self.ships_to_exclude))][self.features]
        X_cat = self.encoder_x.transform(X_pred[cat_features])
        X_num = X_pred[num_features]
        self.X_enc_pred = np.hstack([X_cat, X_num])
        self.log.debug(f"Shapes: X {X.shape}, X_enc_train {self.X_enc_train.shape}, X_enc_pred {self.X_enc_pred.shape}")
        
        # Transform y
        y = self.df[(self.df['isTrain']==True)&(~self.df['Last_Future_ship_to'].isin(self.ships_to_exclude))][target]
        self.encoder_y = LabelEncoder()
        # y is a DataFrame, converting to 1D array
        self.y_enc_train = self.encoder_y.fit_transform(y.values.ravel())
        self.log.debug(f"Shape: y_enc_train {self.y_enc_train.shape}")
        
    def validate(self):
        """ Split the dataset into the training and the validation parts for training and validation """
        self.get_encoded()
        # Training-Validation split
        X_train, X_valid, y_train, y_valid = train_test_split(self.X_enc_train, self.y_enc_train,
            test_size=0.3, random_state=42, stratify=self.y_enc_train)
        # Training, validation, Cross-Validation
        t0 = time()
        self.model.fit(X_train, y_train)
        self.log.debug(f"Training finished in {time() - t0:.3f} sec.")
        t0 = time()
        y_pred = self.model.predict(X_valid)
        self.val_score = balanced_accuracy_score(y_valid, y_pred)
        self.log.info(f"Balanced accuracy score: {self.val_score:.3f}")
        self.log.debug(f"Validation finished in {time() - t0:.1f} sec.")
        t0 = time()
        val_cv_score = cross_val_score(self.model, self.X_enc_train, self.y_enc_train, cv=3, scoring='balanced_accuracy')
        self.val_cv_score = np.array([round(item, 5) for item in val_cv_score])
        self.log.info(f"Cross-validation average score: {self.val_cv_score.mean():.3f}")
        self.log.debug(f"Cross-validation finished in {time() - t0:.3f} sec.")
        # print statistics
        self.get_statistics(X_valid, y_valid)

    def get_statistics(self, X_valid, y_valid):
        """ Print statistics """
        self.find_top_3(X_valid)
        # y_valid закодирован, поэтому инвертируем как было
        self.proba['y_valid'] = self.encoder_y.inverse_transform(y_valid)
        self.proba['correct_1'] = self.proba.apply(lambda x: int(x.top_1_class==x.y_valid),axis=1)
        self.proba['correct_2'] = self.proba.apply(lambda x: int(x.top_2_class==x.y_valid),axis=1)
        self.proba['correct_3'] = self.proba.apply(lambda x: int(x.top_3_class==x.y_valid),axis=1)

        # Оставим только новые столбцы с информацией по 3-м классам с наивысшей уверенностью
        self.proba = self.proba.loc[:,'top_1_class':]

        # Всего предсказаний
        total = self.proba.shape[0]
        # Количество верных предсказаний по классам 
        corr_cl1 = self.proba[self.proba.top_1_class==self.proba.y_valid].shape[0]
        corr_cl2 = self.proba[self.proba.top_2_class==self.proba.y_valid].shape[0]
        corr_cl3 = self.proba[self.proba.top_3_class==self.proba.y_valid].shape[0]
        not_correct = total - (corr_cl1 + corr_cl2 + corr_cl3)
        self.log.info(f"""
        Всего предсказаний: {total}
        Правильных предсказаний: {corr_cl1/total*100:.1f}% ({corr_cl1})
        Предсказанных во втором варианте: {corr_cl2/total*100:.2f}% ({corr_cl2})
        Предсказанных в третьем варианте: {corr_cl3/total*100:.3f}% ({corr_cl3})
        Не предсказанных вообще {not_correct/total*100:.3f}% ({not_correct})
        """)

        def get_proba_info(class_num, proba_from, proba_to):
            """
            Получает название класса и возвращает количество правильных, не правильных ответов, а также интервал
            
            """
            correct_num = self.proba[(self.proba[class_num+'_class']==self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            incorrect_num = self.proba[(self.proba[class_num+'_class']!=self.proba.y_valid)& \
                (self.proba[class_num+'_proba']>proba_from)&(self.proba[class_num+'_proba']<=proba_to)].shape[0]
            return correct_num, incorrect_num, (proba_from, proba_to)

        correct, incorrect, index = [], [], []
        for edge in range(20,100,10):
            cor, inc, ind = get_proba_info('top_1',edge/100,(edge+10)/100)
            correct.append(cor)
            incorrect.append(inc)
            index.append(ind)
            
        a = pd.DataFrame(data={'correct':correct[::-1], 'incorrect':incorrect[::-1]}, index=index[::-1])    

        # Напечатать интервальную серию с выводом информации об отношении количеств элементов в интервалах 
        top = 1
        bottom = 0
        rep_list = []
        rep_list.append(f"\n{'Интервал':>12} {'Верных':>8} {'Неверных':>10} {'Нев./Общ.':>11}\n")
        for i in range(len(a)):
            mid = a.index[i][0]
            s = f"{str(a.index[i]):>12} {a.correct[i]:>8} {a.incorrect[i]:>10}"
            if i==0:
                v = a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100
            else:
                v = a.incorrect[:i+1].sum() / (a.incorrect[:i+1].sum() + a.correct[:i+1].sum()) * 100
            rep_list.append("{0} {1:>10.2f} | интервал ({2}, {3}] кол. ошибок / общему кол. предсказаний = {4:.2f}%\n" \
                    .format(s, a.incorrect[i] / (a.incorrect[i] + a.correct[i]) * 100, mid, top, v))
        self.log.info(''.join(rep_list))

    def find_top_3(self, X_valid):
        """ Define top 3 classes for each outlet without an answer """
        y_pred_proba = self.model.predict_proba(X_valid)
        self.proba = pd.DataFrame(data=y_pred_proba, columns=self.model.classes_)
        self.log.debug(f"proba.shape {self.proba.shape}")

        def get_max_3_classes(row):
            """
            Получает серию из предсказаний размерностью n-классов и возвращает три класса с максимальной вероятностью
            и значания вероятности для этих классов.
            
            """
            ser = pd.Series(data=row.values, index=self.model.classes_)
            ser.sort_values(inplace=True, ascending=False)
            return ser[0:3].index[0],ser[0:3].values[0], \
                ser[0:3].index[1],ser[0:3].values[1], \
                ser[0:3].index[2],ser[0:3].values[2]

        self.proba['top_1_class'], self.proba['top_1_proba'], \
            self.proba['top_2_class'], self.proba['top_2_proba'], \
            self.proba['top_3_class'], self.proba['top_3_proba'] = zip(*self.proba.apply(get_max_3_classes, axis=1))

        self.proba['top_1_class'] = self.encoder_y.inverse_transform(self.proba['top_1_class'].values.ravel())
        self.proba['top_2_class'] = self.encoder_y.inverse_transform(self.proba['top_2_class'].values.ravel())
        self.proba['top_3_class'] = self.encoder_y.inverse_transform(self.proba['top_3_class'].values.ravel())
        
    def fit(self):
        """ Training on full data set """
        self.log.info("Final training...")
        self.get_encoded()
        t0 = time()
        self.model.fit(self.X_enc_train, self.y_enc_train)
        self.log.debug(f"Final training finished in {time() - t0:.3f} sec.")

    def get_report(self):
        """ Prepare a new report """

        self.log.info("Calculate proba...")
        t0 = time()
        # Generate proba
        self.find_top_3(self.X_enc_pred)
        X_pred = self.df[(self.df['isTrain']==False)& \
                         (~self.df['Last_Future_ship_to'] \
                          .isin(self.ships_to_exclude))][self.features]
        X_pred.reset_index(inplace=True)
        df_concat = pd.concat([X_pred['SWE_Store_Key'], self.proba.loc[:,'top_1_class':]], axis=1,join='inner')
        df_info = self.df_codes.merge(right=df_concat,how='left',on='SWE_Store_Key')
        df_info['SWE_Store_Key'] = df_info['SWE_Store_Key'].astype('str')
        self.log.debug(f"Done in {time() - t0:.3f} sec.")

        self.log.info("Open report...")
        t0 = time()
        self.workbook = openpyxl.load_workbook(self.report_file)
        self.log.debug(f"Done in {time() - t0:.3f} sec.")
        
        self.log.info("Format report...")
        t0 = time()
        worksheet = self.workbook['Sheet1']
        rows = dataframe_to_rows(df_info, index=False, header=True)
        proba_row = 2
        proba_col = 54    # BB column
        target_col = 51   # AY column
        region_col = 49
        name_col = 49
        # Setup column width, setup title text, font and alignment
        widths = [19,11,5,11,5,11,5]
        captions = ['SWE Store Key','1 class',' 1 proba','2 class','2 proba','3 class','3 proba']
        for i in range(7):
            worksheet.column_dimensions[get_column_letter(proba_col + i)].width = widths[i]
            cell = get_column_letter(proba_col + i) + str(proba_row)
            worksheet[cell] = captions[i]
            worksheet[cell].font = Font(name='Times New Roman', size=12, bold=True)
            worksheet[cell].alignment = Alignment(horizontal='left', vertical='top')
            worksheet[cell].fill = PatternFill("solid", fgColor="00CCFFCC")
        # Go through all rows
        for r_idx, row in enumerate(rows, proba_row):
            # If proba is defined or the very first row with a title
            if type(row[2])==float and not pd.isnull(row[2]):
                # Go through columns in a row
                for c_idx, value in enumerate(row, proba_col):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
                    worksheet.cell(row=r_idx, column=c_idx).font = Font(name='Arial', size=10, bold=False)
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='left', vertical='top')
                if r_idx > 2:
                    proba = float(row[2])
                    if proba >= 0.9:
                        worksheet.cell(row=r_idx, column=target_col, value=row[1])
                        worksheet.cell(row=r_idx, column=target_col).fill = PatternFill("solid", fgColor="00CCFFCC")
                    elif proba >= 0.7:
                        worksheet.cell(row=r_idx, column=target_col, value=row[1])
                        worksheet.cell(row=r_idx, column=target_col).fill = PatternFill("solid", fgColor="00FFCC99")
                    else:
                        worksheet.cell(row=r_idx, column=target_col, value=row[1])
                        worksheet.cell(row=r_idx, column=target_col).fill = PatternFill("solid", fgColor="00FF9900")
        self.log.debug(f"Done in {time() - t0:.3f} sec.")

    def save_report(self):
        self.log.info("Save output file...")
        t0 = time()
        self.workbook.save(self.output_file)
        self.log.debug(f"Saved in {time() - t0:.3f} sec.")
        self.log.info(f"New report saved as '{self.output_file}''")